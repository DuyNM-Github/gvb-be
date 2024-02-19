from rest_framework.response import Response
from django.http import FileResponse
from rest_framework.decorators import api_view
from msoffice2pdf import convert as doc_convert
import os, pathlib, zipfile, shutil, uuid
from .serializers import UploadedFileSerializer
from base.models import UploadedFile
from openpyxl import load_workbook
import datetime, shutil


BASE_TEMP = "./temp/"

# Create your views here.
@api_view(['POST'])
def receive_upload_files(request):
    if len(request.FILES) < 1:
        return Response(status=400)

    if 'session' not in request.data.keys():
        session_id = uuid.uuid4()
    else:
        if request.data['session'] is not None:
            session_id = request.data['session']
        else:
            session_id = uuid.uuid4()

    uploaded_files = []
    session_temp = os.path.join(BASE_TEMP, str(session_id))
    if request.FILES is not None and len(request.FILES) >= 1:
        for file_obj in request.FILES.values():
                uploaded_files.append(file_obj.name)    
                chunks = file_obj.chunks(1000000)

                # Save the file to temp folder
                pathlib.Path(session_temp).mkdir(parents=True, exist_ok=True)
                file_path = os.path.join(session_temp, file_obj.name)
                with open(file_path, "wb") as f:
                    for chunk in chunks:
                        f.write(chunk)
                
                # Save the file to database
                uploaded_file = UploadedFile(session=session_id, file_id=uuid.uuid4(), file_name=file_obj.name)
                uploaded_file.save()

    saved_files = UploadedFile.objects.filter(session=session_id)
    serializer = UploadedFileSerializer(saved_files, many=True)
    del saved_files
    
    return Response(data={"session": session_id, "uploaded_files": serializer.data})


@api_view(['POST'])
def process_uploaded_files(request):
    try:
        ref_id = {}
        if 'company_id' in request.data.keys() and 'client_id' in request.data.keys():
            ref_id['company'] = request.data['company_id']
            ref_id['client'] = request.data['client_id']
        else:
            ref_id['company'] = "unknown"
            ref_id['client'] = "unknown"
    
        data = {}
        data['input'] = request.data
        if 'session' not in request.data.keys():
            return Response(status=400)
        if request.data['session'] is None:
            return Response(status=400)
        
        session_folder = os.path.join(BASE_TEMP, request.data['session'])
        
        if os.path.exists(session_folder) is False:
            return Response(status=404)
        
        # File renaming
        used_files: list = rename_files(session_folder, request, ref_id)
        
        # File conversion
        renamed_files = [filepath for filepath in pathlib.Path(session_folder).iterdir()]
        excel_file = None
        for file in renamed_files:
            if ("doc" in file.name or "docx" in file.name) and ("ANS" in file.name or "BES" in file.name):
                print(f"Converting file {file.name} to PDF")
                doc_convert(source=file, output_dir=session_folder, soft=1)
            if "xlsx" in file.name:
                excel_file = file
        temp_pdfs = [filepath for filepath in pathlib.Path(session_folder).iterdir()]
        for file in temp_pdfs:
            if "pdf" in file.name and ("ANS" in file.name or "BES" in file.name):
                os.rename(file, os.path.join(session_folder, file.name[20:]))
                used_files.append(file.name[20:])
        del renamed_files, temp_pdfs

        # Filling excel sheet
        if excel_file is None:
            return Response(status=404)
        print("Start filling excel sheets")
        filling_excel_sheet(session_folder, excel_file, ref_id)
        del excel_file

        # Zip file
        zip_path = os.path.join(session_folder, f"{ref_id['client']}.zip")
        if os.path.exists(zip_path):
            print("remove zip file if exists")
            os.remove(zip_path)
        processed_files = [filepath for filepath in pathlib.Path(session_folder).iterdir()]
        with zipfile.ZipFile(zip_path, mode='w') as archive:
            for filepath in processed_files:
                try:
                    if filepath.name in used_files:
                        archive.write(filepath, arcname=filepath.name)
                except Exception:
                    print("error occurred when zipping files")
                    pass
        del processed_files, used_files

        response =  FileResponse(
                open(zip_path, 'rb'),
                as_attachment=True,
                filename=(f"{ref_id['client']}.zip")
            )
        shutil.rmtree(session_folder)

        return response
    # return Response(data=data)
    except Exception as e:
        print("Process operation failed")
        print(e)
        session_folder = os.path.join(BASE_TEMP, request.data['session'])
        shutil.rmtree(session_folder)
        return Response(status=500)

@api_view(['DELETE'])
def clear_temp(request):
    for filename in os.listdir(BASE_TEMP):
        file_path = os.path.join(BASE_TEMP, filename)
        try:
            if os.path.isfile(file_path) or os.path.islink(file_path):
                os.unlink(file_path)
            elif os.path.isdir(file_path):
                shutil.rmtree(file_path)
        except Exception as e:
            print('Failed to delete %s. Reason: %s' % (file_path, e))
    return Response()


def rename_files(session_folder, request, ref_id):
    uploaded_files = [filepath for filepath in pathlib.Path(session_folder).iterdir()]
    used_files = []
    process_date = str(datetime.datetime.now().strftime("%Y-%m-%d"))
    for payload in request.data['file']:
        if len(payload['list']) >= 1:
            match payload['name'].strip():
                case "Anschreiben muss Ueberm und BSW enthalten und auf .doc(x) enden":
                    temp_file = [file for file in uploaded_files if file.name == payload['list'][0]['name']][0]
                    os.rename(
                        os.path.join(session_folder, temp_file.name), 
                        os.path.join(session_folder, f"{ref_id['client']}-{ref_id['company']}-Anschreiben.docx")
                    )
                    used_files.append(f"{ref_id['client']}-{ref_id['company']}-Anschreiben.docx")
                    del temp_file
                case "Excel-Datei (muss .xlsx heißen)":
                    temp_file = [file for file in uploaded_files if file.name == payload['list'][0]['name']][0]
                    os.rename(
                        os.path.join(session_folder, temp_file.name), 
                        os.path.join(session_folder, f"{ref_id['company']}-{ref_id['client']}-Beauftragung_Dienstleistung_BSE_Allgemein_{ref_id['company']}.xlsx")
                    )
                    used_files.append(f"{ref_id['company']}-{ref_id['client']}-Beauftragung_Dienstleistung_BSE_Allgemein_{ref_id['company']}.xlsx")
                    del temp_file
                case "Bescheidserwiderung (muss *-BSW-• enthalten)":
                    temp_file = [file for file in uploaded_files if file.name == payload['list'][0]['name']][0]
                    os.rename(
                        os.path.join(session_folder, temp_file.name), 
                        os.path.join(session_folder, f"{ref_id['client']}-{ref_id['company']}-BSW-{process_date}.docx")
                    )
                    used_files.append(f"{ref_id['client']}-{ref_id['company']}-BSW-{process_date}.docx")
                    del temp_file
                case "Neue Patentansprüche Korrekturexemplar (muss -neue-ANS- und -korr enthalten)":
                    input_files = [file['name'] for file in payload['list']]
                    temp_files = [file for file in uploaded_files if file.name in input_files]
                    if len(temp_files) > 1:
                        hilfs_count = 1
                        for file in temp_files:
                            if "haupt" in file.name.lower():
                                os.rename (
                                    os.path.join(session_folder, file.name), 
                                    os.path.join(session_folder, f"{ref_id['client']}-{ref_id['company']}-neue-ANS-Haupt-{process_date}_korr.docx")
                                )
                                used_files.append(f"{ref_id['client']}-{ref_id['company']}-neue-ANS-Haupt-{process_date}_korr.docx")
                            elif "hilfs" in file.name.lower():
                                os.rename (
                                    os.path.join(session_folder, file.name), 
                                    os.path.join(session_folder, f"{ref_id['client']}-{ref_id['company']}-neue-ANS-Hilfs{hilfs_count}-{process_date}_korr.docx")
                                )
                                used_files.append(f"{ref_id['client']}-{ref_id['company']}-neue-ANS-Hilfs{hilfs_count}-{process_date}_korr.docx")
                                hilfs_count += 1
                    elif len(temp_files) == 1:
                        file = temp_files[0]
                        if "haupt" in file.name.lower():
                            os.rename (
                                os.path.join(session_folder, file.name), 
                                os.path.join(session_folder, f"{ref_id['client']}-{ref_id['company']}-neue-ANS-Haupt-{process_date}_korr.docx")
                            )
                            used_files.append(f"{ref_id['client']}-{ref_id['company']}-neue-ANS-Haupt-{process_date}_korr.docx")
                        elif "hilfs" in file.name.lower():
                            os.rename (
                                os.path.join(session_folder, file.name), 
                                os.path.join(session_folder, f"{ref_id['client']}-{ref_id['company']}-neue-ANS-Hilfs-{process_date}_korr.docx")
                            )
                            used_files.append(f"{ref_id['client']}-{ref_id['company']}-neue-ANS-Hilfs-{process_date}_korr.docx")
                    del input_files, temp_files
                case "Neue Patentansprüche Reinschrift (muss -neue-ANS- und -rein enthalten )":
                    input_files = [file['name'] for file in payload['list']]
                    temp_files = [file for file in uploaded_files if file.name in input_files]
                    if len(temp_files) > 1:
                        hilfs_count = 1
                        for file in temp_files:
                            if "haupt" in file.name.lower():
                                os.rename (
                                    os.path.join(session_folder, file.name), 
                                    os.path.join(session_folder, f"{ref_id['client']}-{ref_id['company']}-neue-ANS-Haupt-{process_date}_rein.docx")
                                )
                                used_files.append(f"{ref_id['client']}-{ref_id['company']}-neue-ANS-Haupt-{process_date}_rein.docx")
                            elif "hilfs" in file.name.lower():
                                os.rename (
                                    os.path.join(session_folder, file.name), 
                                    os.path.join(session_folder, f"{ref_id['client']}-{ref_id['company']}-neue-ANS-Hilfs{hilfs_count}-{process_date}_rein.docx")
                                )
                                used_files.append(f"{ref_id['client']}-{ref_id['company']}-neue-ANS-Hilfs{hilfs_count}-{process_date}_rein.docx")
                                hilfs_count += 1
                    elif len(temp_files) == 1:
                        file = temp_files[0]
                        if "haupt" in file.name.lower():
                            os.rename (
                                os.path.join(session_folder, file.name), 
                                os.path.join(session_folder, f"{ref_id['client']}-{ref_id['company']}-neue-ANS-Haupt-{process_date}_rein.docx")
                            )
                            used_files.append(f"{ref_id['client']}-{ref_id['company']}-neue-ANS-Haupt-{process_date}_rein.docx")
                        elif "hilfs" in file.name.lower():
                            os.rename (
                                os.path.join(session_folder, file.name), 
                                os.path.join(session_folder, f"{ref_id['client']}-{ref_id['company']}-neue-ANS-Hilfs-{process_date}_rein.docx")
                            )
                            used_files.append(f"{ref_id['client']}-{ref_id['company']}-neue-ANS-Hilfs-{process_date}_rein.docx")
                    del input_files, temp_files
                case "Neue Beschreibungsseiten (muss -neue-BES- und -korr enthalten)":
                    input_files = [file['name'] for file in payload['list']]
                    temp_files = [file for file in uploaded_files if file.name in input_files]
                    if len(temp_files) > 1:
                        hilfs_count = 1
                        for file in temp_files:
                            if "haupt" in file.name.lower():
                                os.rename (
                                    os.path.join(session_folder, file.name), 
                                    os.path.join(session_folder, f"{ref_id['client']}-{ref_id['company']}-neue-BES-Haupt-{process_date}_korr.docx")
                                )
                                used_files.append(f"{ref_id['client']}-{ref_id['company']}-neue-BES-Haupt-{process_date}_korr.docx")
                            elif "hilfs" in file.name.lower():
                                os.rename (
                                    os.path.join(session_folder, file.name), 
                                    os.path.join(session_folder, f"{ref_id['client']}-{ref_id['company']}-neue-BES-Hilfs{hilfs_count}-{process_date}_korr.docx")
                                )
                                used_files.append(f"{ref_id['client']}-{ref_id['company']}-neue-BES-Hilfs{hilfs_count}-{process_date}_korr.docx")
                                hilfs_count += 1
                    elif len(temp_files) == 1:
                        file = temp_files[0]
                        if "haupt" in file.name.lower():
                            os.rename (
                                os.path.join(session_folder, file.name), 
                                os.path.join(session_folder, f"{ref_id['client']}-{ref_id['company']}-neue-BES-Haupt-{process_date}_korr.docx")
                            )
                            used_files.append(f"{ref_id['client']}-{ref_id['company']}-neue-BES-Haupt-{process_date}_korr.docx")
                        elif "hilfs" in file.name.lower():
                            os.rename (
                                os.path.join(session_folder, file.name), 
                                os.path.join(session_folder, f"{ref_id['client']}-{ref_id['company']}-neue-BES-Hilfs-{process_date}_korr.docx")
                            )
                            used_files.append(f"{ref_id['client']}-{ref_id['company']}-neue-BES-Hilfs-{process_date}_korr.docx")
                        del input_files, temp_files
                case "Neue Beschreibungsseiten (muss -neue-BES- und -rein enthalten)":
                    input_files = [file['name'] for file in payload['list']]
                    temp_files = [file for file in uploaded_files if file.name in input_files]
                    if len(temp_files) > 1:
                        hilfs_count = 1
                        for file in temp_files:
                            if "haupt" in file.name.lower():
                                os.rename (
                                    os.path.join(session_folder, file.name), 
                                    os.path.join(session_folder, f"{ref_id['client']}-{ref_id['company']}-neue-BES-Haupt-{process_date}_rein.docx")
                                )
                                used_files.append(f"{ref_id['client']}-{ref_id['company']}-neue-BES-Haupt-{process_date}_rein.docx")
                            elif "hilfs" in file.name.lower():
                                os.rename (
                                    os.path.join(session_folder, file.name), 
                                    os.path.join(session_folder, f"{ref_id['client']}-{ref_id['company']}-neue-BES-Hilfs{hilfs_count}-{process_date}_rein.docx")
                                )
                                used_files.append(f"{ref_id['client']}-{ref_id['company']}-neue-BES-Hilfs{hilfs_count}-{process_date}_rein.docx")
                                hilfs_count += 1
                    elif len(temp_files) == 1:
                        file = temp_files[0]
                        if "haupt" in file.name.lower():
                            os.rename (
                                os.path.join(session_folder, file.name), 
                                os.path.join(session_folder, f"{ref_id['client']}-{ref_id['company']}-neue-BES-Haupt-{process_date}_rein.docx")
                            )
                            used_files.append(f"{ref_id['client']}-{ref_id['company']}-neue-BES-Haupt-{process_date}_rein.docx")
                        elif "hilfs" in file.name.lower():
                            os.rename (
                                os.path.join(session_folder, file.name), 
                                os.path.join(session_folder, f"{ref_id['client']}-{ref_id['company']}-neue-BES-Hilfs-{process_date}_rein.docx")
                            )
                            used_files.append(f"{ref_id['client']}-{ref_id['company']}-neue-BES-Hilfs-{process_date}_rein.docx")
                    del input_files, temp_files
                case "Optionale Zeichnungen":
                    temp_file = [file for file in uploaded_files if file.name == payload['list'][0]['name']][0]
                    extension = temp_file.name.split(".")[1]
                    os.rename(
                        os.path.join(session_folder, temp_file.name), 
                        os.path.join(session_folder, f"{ref_id['client']}-{ref_id['company']}-Reinzeichnungen.{extension}")
                    )
                    used_files.append(f"{ref_id['client']}-{ref_id['company']}-Reinzeichnungen.{extension}")
                    del temp_file, extension
                case "Rechnung (muss *Rechnung* heißen und auf .pdf enden)":
                    temp_file = [file for file in uploaded_files if file.name == payload['list'][0]['name']][0]
                    os.rename(
                        os.path.join(session_folder, temp_file.name), 
                        os.path.join(session_folder, f"{ref_id['client']}-{ref_id['company']}-Rechnung-{process_date}.pdf")
                    )
                    used_files.append(f"{ref_id['client']}-{ref_id['company']}-Rechnung-{process_date}.pdf")
                    del temp_file
    del uploaded_files
    return used_files
    

def filling_excel_sheet(session_folder, excel_file, ref_id):
    post_process_files = [file for file in pathlib.Path(session_folder).iterdir()]
    wb = load_workbook(excel_file)
    ws = wb.active
    for row in ws.iter_rows(min_row=3, max_col=3, max_row=23):
        # print(f"{row[0].value} - {row[1].value} - {row[2].value}")
        match (row[0].value.strip(), row[2].value.strip()):
            case ("Aktenzeichen des Auftragnehmers", _):
                row[1].value = ref_id['company']
            case ("Anschreiben/Kommentar", _):
                files = [file.name for file in post_process_files if "Anschreiben" in file.name]
                if len(files) > 0:
                    row[1].value = files[0]
                    del files
            case ("Entwurf für Bescheidserwiderung/Kommentar", _):
                files = [file.name for file in post_process_files if "BSW" in file.name]
                if len(files) > 0:
                    row[1].value = files[0]
                    del files
            case ("Geänderte Ansprüche (Reinschrift)", "DOCX"):
                files = [file.name for file in post_process_files if 
                         "ANS" in file.name 
                         and "_rein" in file.name
                         and "docx" in file.name]
                if len(files) == 1:
                    row[1].value = files[0]
                elif len(files) > 1:
                    sorted_file = []
                    for file in files:
                        if "Haupt" in file:
                            sorted_file.append(file)
                    for file in files:
                        if "Hilfs" in file:
                            sorted_file.append(file)
                    row[1].value = "//".join(sorted_file)
                del files
            case ("Geänderte Ansprüche (mit Anmerkungen)", "DOCX"):
                files = [file.name for file in post_process_files if 
                         "ANS" in file.name 
                         and "_korr" in file.name
                         and "docx" in file.name]
                if len(files) == 1:
                    row[1].value = files[0]
                elif len(files) > 1:
                    sorted_file = []
                    for file in files:
                        if "Haupt" in file:
                            sorted_file.append(file)
                    for file in files:
                        if "Hilfs" in file:
                            sorted_file.append(file)
                    row[1].value = "//".join(sorted_file)
                del files
            case ("Geänderte Ansprüche (Reinschrift)", "PDFA/ANNEXF"):
                files = [file.name for file in post_process_files if 
                         "ANS" in file.name 
                         and "_rein" in file.name
                         and "pdf" in file.name]
                if len(files) == 1:
                    row[1].value = files[0]
                elif len(files) > 1:
                    sorted_file = []
                    for file in files:
                        if "Haupt" in file:
                            sorted_file.append(file)
                    for file in files:
                        if "Hilfs" in file:
                            sorted_file.append(file)
                    row[1].value = "//".join(sorted_file)
                del files
            case ("Geänderte Ansprüche (mit Anmerkungen)", "PDFA/ANNEXF"):
                files = [file.name for file in post_process_files if 
                         "ANS" in file.name 
                         and "_korr" in file.name
                         and "pdf" in file.name]
                if len(files) == 1:
                    row[1].value = files[0]
                elif len(files) > 1:
                    sorted_file = []
                    for file in files:
                        if "Haupt" in file:
                            sorted_file.append(file)
                    for file in files:
                        if "Hilfs" in file:
                            sorted_file.append(file)
                    row[1].value = "//".join(sorted_file)
                del files
            case ("Geänderte Beschreibungsseiten (Reinschrift)", "DOCX"):
                files = [file.name for file in post_process_files if 
                         "BES" in file.name 
                         and "_rein" in file.name
                         and "docx" in file.name]
                if len(files) == 1:
                    row[1].value = files[0]
                elif len(files) > 1:
                    sorted_file = []
                    for file in files:
                        if "Haupt" in file:
                            sorted_file.append(file)
                    for file in files:
                        if "Hilfs" in file:
                            sorted_file.append(file)
                    row[1].value = "//".join(sorted_file)
                del files
            case ("Geänderte Beschreibungsseiten (mit Anmerkungen)", "DOCX"):
                files = [file.name for file in post_process_files if 
                         "BES" in file.name 
                         and "_korr" in file.name
                         and "docx" in file.name]
                if len(files) == 1:
                    row[1].value = files[0]
                elif len(files) > 1:
                    sorted_file = []
                    for file in files:
                        if "Haupt" in file:
                            sorted_file.append(file)
                    for file in files:
                        if "Hilfs" in file:
                            sorted_file.append(file)
                    row[1].value = "//".join(sorted_file)
                del files
            case ("Geänderte Beschreibungsseiten (Reinschrift)", "PDFA/ANNEXF"):
                files = [file.name for file in post_process_files if 
                         "BES" in file.name 
                         and "_rein" in file.name
                         and "pdf" in file.name]
                if len(files) == 1:
                    row[1].value = files[0]
                elif len(files) > 1:
                    sorted_file = []
                    for file in files:
                        if "Haupt" in file:
                            sorted_file.append(file)
                    for file in files:
                        if "Hilfs" in file:
                            sorted_file.append(file)
                    row[1].value = "//".join(sorted_file)
                del files
            case ("Geänderte Beschreibungsseiten (mit Anmerkungen)", "PDFA/ANNEXF"):
                files = [file.name for file in post_process_files if 
                         "BES" in file.name 
                         and "_korr" in file.name
                         and "pdf" in file.name]
                if len(files) == 1:
                    row[1].value = files[0]
                elif len(files) > 1:
                    sorted_file = []
                    for file in files:
                        if "Haupt" in file:
                            sorted_file.append(file)
                    for file in files:
                        if "Hilfs" in file:
                            sorted_file.append(file)
                    row[1].value = "//".join(sorted_file)
                del files
            case ("Rechnungsdatei", "PDFA"):
                files = [file.name for file in post_process_files if "Rechnung" in file.name]
                if len(files) > 0:
                    row[1].value = files[0]
                    del files
            case _:
                print(f"something happened - {(row[0].value, row[2].value)}")
                pass
    wb.save(excel_file)
    wb.close()
    del post_process_files
    